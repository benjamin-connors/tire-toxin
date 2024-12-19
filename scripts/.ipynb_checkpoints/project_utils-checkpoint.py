import os
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
import numpy as np
import sys

# Default decimal formatting
DEFAULT_DECIMALS = {
    'Differential Pressure (kPa)': 3, 
    'Absolute Pressure (kPa)': 3,
    'Temperature (°C)': 2,
    'Water Level (m)': 5,
    'Barometric Pressure (kPa)': 3
}

def format_and_save_excel(df_corrected, output_file, sheet_name='Sheet1', decimals=None):
    """
    Formats and saves a DataFrame to an Excel file with OpenPyxl, applying number formatting for display.

    Parameters:
    - df_corrected (pd.DataFrame): Data to write to Excel.
    - output_file (Path or str): Path to save the Excel file.
    - sheet_name (str): Name of the sheet in the Excel file.
    - decimals (dict): A dictionary specifying the number of decimal places for each column.
                       If None, defaults to DEFAULT_DECIMALS.
    """
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Use default decimals if none provided
    if decimals is None:
        decimals = DEFAULT_DECIMALS

    # Reset index to include the 'Datetime' column in the DataFrame
    df_corrected_reset = df_corrected.reset_index()
    df_corrected_reset.rename(columns={'index': 'Datetime'}, inplace=True)

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # Write DataFrame rows to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df_corrected_reset, index=False, header=True), start=1):
        worksheet.append(row)

        # Apply header formatting
        if r_idx == 1:
            for cell in worksheet[r_idx]:
                cell.font = Font(bold=True)
    
    # Remove header row borders
    for cell in worksheet[1]:
        cell.border = Border(left=Side(border_style=None), 
                             right=Side(border_style=None), 
                             top=Side(border_style=None), 
                             bottom=Side(border_style=None))

    # Adjust column widths and apply number formatting
    for col_idx, col in enumerate(worksheet.columns, start=1):
        max_length = 0
        column = col[0].column_letter  # Get the column name
        column_name = worksheet.cell(row=1, column=col_idx).value

        # Adjust column widths
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        
        worksheet.column_dimensions[column].width = max_length + 2  # Add padding

        # Apply number formatting if column has a decimal specification
        if column_name in decimals:
            number_format = f"0.{'0' * decimals[column_name]}"
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for data_cell in cell:
                    if isinstance(data_cell.value, (int, float)):
                        data_cell.number_format = number_format

    # Save workbook to file
    workbook.save(output_file)

def determine_file_type_and_site(file):
    """
    Determines the file type (BT or non-BT), the site name, and the associated BT file 
    if applicable. Returns the dataframe loaded from the file, site name, file type, 
    output file path, and BT file path.

    Parameters:
    - file (Path): The path of the file to process.

    Returns:
    - df_corrected (pd.DataFrame): DataFrame loaded from the input file.
    - site_name (str): The site name identified from the file.
    - file_type (str): 'BT' or 'non-BT' indicating the type of file.
    - output_file (Path): The path to the output file based on the site and file type.
    - bt_file (Path or None): The path to the BT file for barometric correction, if applicable.
    """
    # Determine file type (bluetooth / non-bluetooth) and load file accordingly
    first_row = pd.read_excel(file, nrows=1, header=None)

    if "Plot Title" in str(first_row.iloc[0, 0]):
        print("Non-BT File Identified. Reading File")
        colnames = ['Datetime', 'Absolute Pressure (kPa)', 'Temperature (°C)']
        df_corrected = pd.read_excel(file, header=1, index_col=0, parse_dates=True, usecols="B:D", names=colnames)
        file_type = 'non-BT'
    elif "Date-Time" in str(first_row.iloc[0, 1]):
        print("BT File Detected!")
        colnames = ['Datetime', 'Differential Pressure (kPa)', 'Absolute Pressure (kPa)',
                     'Temperature (°C)', 'Water Level (m)', 'Barometric Pressure (kPa)']
        df_corrected = pd.read_excel(file, header=0, index_col=0, parse_dates=True, usecols="B:G", names=colnames)
        file_type = 'BT'
    else:
        raise ValueError("Unknown file format")

    # Identify site
    if ("22084122" in file.name) or ("cat_beaconsBT" in file.name):  # cat_beaconsBT
        site_name = "cat_beaconsBT"
    elif ("22084123" in file.name) or ("northfieldBT" in file.name):  # northfieldBT
        site_name = "northfieldBT"
    elif ("22084124" in file.name) or ("chase_usBT" in file.name):  # chase_usBT
        site_name = "chase_usBT"
    elif ("cat_beacons_" in file.name):  # cat_beacon
        site_name = "cat_beacons"
    elif ("northfield_" in file.name):  # northfield
        site_name = "northfield"
    elif ("chase_us" in file.name):  # chase_us
        site_name = "chase_us"
    elif ("chase_ds" in file.name):  # chase_ds
        site_name = "chase_ds"
    else:
        raise ValueError("Unknown site")

    # Generate the output file path
    output_directory = file.parents[2] / 'processed'
    output_file = output_directory / f"{site_name}_stage_master.xlsx"

    # Determine the BT file path (if applicable)
    if site_name in ['northfield', 'cat_beacons', 'chase_us']:
        bt_file = output_directory / f"{site_name}BT_stage_master.xlsx"
    elif site_name == "chase_ds":
        bt_file = output_directory / "chase_usBT_stage_master.xlsx"
    else:
        bt_file = None

    return df_corrected, site_name, file_type, output_file, bt_file


def calculate_differential_pressure(df, df_baro):
    """
    Calculates Differential Pressure (kPa) based on Absolute Pressure and Barometric Pressure.

    Parameters:
    - df (pd.DataFrame): DataFrame containing 'Absolute Pressure (kPa)' and 'Differential Pressure (kPa)' columns.
    - df_baro (pd.DataFrame): DataFrame containing the barometric presssure data for the correction in a 'Barometric Pressure (kPa)' column.

    Returns:
    - df (pd.DataFrame): DataFrame with 'Differential Pressure (kPa)' filled for missing values.
    - corrected_baro_count (int): Number of corrections made.
    """
    
    corrected_baro_count = 0
    failed_baro_count = 0

    # Apply correction only to rows where 'Differential Pressure (kPa)' is missing
    if 'Differential Pressure (kPa)' in df.columns:
        missing_data_mask = df['Differential Pressure (kPa)'].isna()

        # If there are rows with missing data
        if missing_data_mask.any():
            for idx, row in df[missing_data_mask].iterrows():
                # Calculate the differential pressure for missing values
                if pd.isna(row['Differential Pressure (kPa)']) and pd.notna(row['Absolute Pressure (kPa)']):
                    # Calculate the time difference between the current row and all barometric timestamps
                    time_diff = np.abs(df_baro.index - pd.to_datetime(row.name))

                    # Filter for valid timestamps within the 10-minute window
                    valid_times = time_diff[time_diff <= pd.Timedelta(minutes=10)]

                    if not valid_times.empty:
                        # Get the index (datetime) corresponding to the minimum time difference
                        nearest_time = time_diff[time_diff == valid_times.min()].index[0]  # This gives the actual datetime index

                        # Get the barometric pressure at the nearest timestamp
                        barometric_pressure = df_baro.loc[nearest_time, 'Barometric Pressure (kPa)']

                        # Calculate Differential Pressure
                        diff_pressure = row['Absolute Pressure (kPa)'] - barometric_pressure
                        df.at[idx, 'Differential Pressure (kPa)'] = diff_pressure
                        corrected_baro_count += 1

                    else:
                        failed_baro_count += 1

    return df, corrected_baro_count, failed_baro_count

def calculate_water_level(df, g=9.81):
    """
    Calculates Water Level (m) from Differential Pressure (kPa) and temperature.
    
    Parameters:
    - df (pd.DataFrame): DataFrame containing 'Differential Pressure (kPa)' and 'Temperature (°C)' columns.
    - g (float): Gravitational acceleration (m/s^2), default is 9.81 m/s^2.
    
    Returns:
    - df (pd.DataFrame): DataFrame with 'Water Level (m)' filled for missing values.
    - corrected_water_level_count (int): Number of corrections made.
    """
    corrected_water_level_count = 0

    # Apply correction only to rows where 'Water Level (m)' is missing
    if 'Water Level (m)' in df.columns:
        missing_data_mask = df['Water Level (m)'].isna()

        # If there are rows with missing data
        if missing_data_mask.any():
            for idx, row in df[missing_data_mask].iterrows():
                # Calculate the water level for missing values
                if pd.isna(row['Water Level (m)']) and pd.notna(row['Differential Pressure (kPa)']):
                    T = row['Temperature (°C)']
                    rho = 999.84 - 0.067 * T  # Simplified density formula for freshwater (kg/m³)
                    df.at[idx, 'Water Level (m)'] = (row['Differential Pressure (kPa)'] * 1000) / (rho * g)
                    corrected_water_level_count += 1

    return df, corrected_water_level_count


