import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter
import gspread
from config import credentials
from openpyxl.styles import Font, Side, Border

# Default decimal formatting
DEFAULT_DECIMALS = {
    'Differential Pressure (kPa)': 3, 
    'Absolute Pressure (kPa)': 3,
    'Temperature (°C)': 2,
    'Water Level (m)': 5,
    'Barometric Pressure (kPa)': 3
}

def save_formatted_stage_file(df_corrected, output_file, sheet_name='Sheet1', decimals=None):
    """
    Formats and saves a DataFrame to an Excel file with OpenPyxl, applying number formatting for display.

    Parameters:
    - df_corrected (pd.DataFrame): Data to write to Excel.
    - output_file (Path or str): Path to save the Excel file.
    - sheet_name (str): Name of the sheet in the Excel file.
    - decimals (dict): A dictionary specifying the number of decimal places for each column.
                       If None, defaults to DEFAULT_DECIMALS.
    """
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Use default decimals if none provided
    if decimals is None:
        decimals = DEFAULT_DECIMALS

    # Reset index to include the 'Datetime' column in the DataFrame
    df_corrected_reset = df_corrected.reset_index()
    df_corrected_reset.rename(columns={'index': 'Datetime'}, inplace=True)

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    # Write DataFrame rows to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df_corrected_reset, index=False, header=True), start=1):
        worksheet.append(row)

        # Apply header formatting
        if r_idx == 1:
            for cell in worksheet[r_idx]:
                cell.font = Font(bold=True)
    
    # Remove header row borders
    for cell in worksheet[1]:
        cell.border = Border(left=Side(border_style=None), 
                             right=Side(border_style=None), 
                             top=Side(border_style=None), 
                             bottom=Side(border_style=None))

    # Adjust column widths and apply number formatting
    for col_idx, col in enumerate(worksheet.columns, start=1):
        max_length = 0
        column = col[0].column_letter  # Get the column name
        column_name = worksheet.cell(row=1, column=col_idx).value

        # Adjust column widths
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        
        worksheet.column_dimensions[column].width = max_length + 2  # Add padding

        # Apply number formatting if column has a decimal specification
        if column_name in decimals:
            number_format = f"0.{'0' * decimals[column_name]}"
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for data_cell in cell:
                    if isinstance(data_cell.value, (int, float)):
                        data_cell.number_format = number_format

    # Save workbook to file
    workbook.save(output_file)

def read_stage_file(file, stats_flag=True):
    # Determine file type (bluetooth / non-bluetooth) and load file accordingly
    first_row = pd.read_excel(file, nrows=1, header=None)
    # Define the column order for master files
    master_cols = ['Differential Pressure (kPa)', 'Absolute Pressure (kPa)', 'Temperature (°C)', 'Barometric Pressure (kPa)', 'Water Level (m)']
    
    # logics for identifying different input file structures
    if "Plot Title" in str(first_row.iloc[0, 0]): # Non-BT
        print("Non-BT File Identified. Reading File")
        # Define the column names you want to read from file and read in data
        colnames = ['Datetime', 'Absolute Pressure (kPa)', 'Temperature (°C)']
        df = pd.read_excel(file, header=1, index_col=0, parse_dates=True, usecols="B:D", names=colnames)
        # Add any missing columns filled with nan
        for col in master_cols:
            if col not in df.columns:
                df[col] = pd.NA
        # Reorder columns to match final column order
        df = df[master_cols]

    elif "Date-Time" in str(first_row.iloc[0, 1]) and "Absolute Pressure , kPa" in str(first_row.iloc[0, 3]): # BT sensor without stats
        print("BT File (no stats) Detected!")
        colnames = ['Datetime', 'Differential Pressure (kPa)', 'Absolute Pressure (kPa)',
                     'Temperature (°C)', 'Water Level (m)', 'Barometric Pressure (kPa)']
        df = pd.read_excel(file, header=0, index_col=0, parse_dates=True, usecols="B:G", names=colnames)
        # Add any missing columns filled with nan
        for col in master_cols:
            if col not in df.columns:
                df[col] = pd.NA
        # Reorder columns to match final column order
        df = df[master_cols]
        
    elif "Date-Time" in str(first_row.iloc[0, 1]) and "Differential Pressure - Max , kPa" in str(first_row.iloc[0, 3]) and stats_flag==True: # BT sensor with stats
        print("BT File (with stats) Detected!")
        colnames = ['Datetime', 'Differential Pressure (kPa)', 'Absolute Pressure (kPa)',
                     'Temperature (°C)', 'Water Level (m)', 'Barometric Pressure (kPa)']
        df = pd.read_excel(file, header=0, index_col=0, parse_dates=True, usecols="B,F,K,P,S,R", names=colnames)
        # Add any missing columns filled with nan
        for col in master_cols:
            if col not in df.columns:
                df[col] = pd.NA
        # Reorder columns to match final column order
        df = df[master_cols]
        
    elif "Date-Time" in str(first_row.iloc[0, 1]) and "Differential Pressure - Max , kPa" in str(first_row.iloc[0, 3]) and stats_flag==False: # BT sensor with stats (ignore stats)
        print("BT File (with stats) Detected!")
        colnames = ['Datetime', 'Differential Pressure (kPa)', 'Absolute Pressure (kPa)',
                     'Temperature (°C)', 'Water Level (m)', 'Barometric Pressure (kPa)']
        df = pd.read_excel(file, header=0, index_col=0, parse_dates=True, usecols="B,C,H,M,S,R", names=colnames)
        # Add any missing columns filled with nan
        for col in master_cols:
            if col not in df.columns:
                df[col] = pd.NA
        # Reorder columns to match final column order
        df = df[master_cols]
    
    else:
        raise ValueError("Unknown file format")

    return df

def read_baro_file(site_name):
    master_directory = Path(r"H:\tire-toxin\data\Discharge\Stage\processed")
    # Determine the baro file path (if applicable)
    if site_name in ['chase_us', 'chase_ds']:
        baro_file = master_directory / f"chase_usBT_stage_master.xlsx"
    elif site_name == "cat_beacons":
        baro_file = master_directory / f"cat_beaconsBT_stage_master.xlsx"
    elif site_name == "northfield_bridge":
        baro_file = master_directory / "northfield_poolBT_stage_master.xlsx"
    else:
        baro_file = None
        df_baro = None

    if baro_file is not None and baro_file.exists():
        df_baro = pd.read_excel(baro_file, header=0, index_col=0, parse_dates=True)
        if df_baro.index.name != 'Datetime':
            df_baro.index.name = 'Datetime'

    return df_baro

def calculate_differential_pressure(df, df_baro):
    """
    Calculates Differential Pressure (kPa) based on Absolute Pressure and Barometric Pressure.

    Parameters:
    - df (pd.DataFrame): DataFrame containing 'Absolute Pressure (kPa)' and 'Differential Pressure (kPa)' columns.
    - df_baro (pd.DataFrame): DataFrame containing the barometric pressure data for the correction in a 'Barometric Pressure (kPa)' column.

    Returns:
    - df (pd.DataFrame): DataFrame with 'Differential Pressure (kPa)' filled for missing values.
    - corrected_baro_count (int): Number of corrections made.
    - failed_baro_count (int): Number of corrections that failed due to no valid data within the threshold.
    """
    corrected_baro_count = 0
    failed_baro_count = 0

    duplicated_rows = df_baro[df_baro.index.duplicated(keep=False)]

    # Sort the duplicated rows by index
    duplicated_rows_sorted = duplicated_rows.sort_index()

    # Export the sorted duplicated rows to an Excel file
    duplicated_rows_sorted.to_excel('duplicated_rows_sorted.xlsx', index=True)

    # Apply correction only to rows where 'Differential Pressure (kPa)' is missing
    if 'Differential Pressure (kPa)' in df.columns:
        missing_data_mask = df['Differential Pressure (kPa)'].isna()

        # If there are rows with missing data
        if missing_data_mask.any():
            for idx, row in df[missing_data_mask].iterrows():
                # Proceed only if Absolute Pressure is available
                if pd.notna(row['Absolute Pressure (kPa)']):
                    row_time = pd.to_datetime(row.name)

                    # Attempt to find the nearest time
                    nearest_idx = df_baro.index.get_indexer([row_time], method='nearest')[0]
                    nearest_time = df_baro.index[nearest_idx]

                    # Check if the nearest time is within the 10-minute threshold
                    if abs(nearest_time - row_time) <= pd.Timedelta(minutes=10):
                        barometric_pressure = df_baro.loc[nearest_time, 'Barometric Pressure (kPa)']

                        # Calculate Differential Pressure
                        diff_pressure = row['Absolute Pressure (kPa)'] - barometric_pressure
                        df.at[idx, 'Differential Pressure (kPa)'] = diff_pressure
                        corrected_baro_count += 1
                    else:
                        # Nearest time is outside the valid range
                        failed_baro_count += 1

    return df, corrected_baro_count, failed_baro_count

def calculate_water_level(df, g=9.81):
    """
    Calculates Water Level (m) from Differential Pressure (kPa) and temperature.
    
    Parameters:
    - df (pd.DataFrame): DataFrame containing 'Differential Pressure (kPa)' and 'Temperature (°C)' columns.
    - g (float): Gravitational acceleration (m/s^2), default is 9.81 m/s^2.
    
    Returns:
    - df (pd.DataFrame): DataFrame with 'Water Level (m)' filled for missing values.
    - corrected_water_level_count (int): Number of corrections made.
    """
    corrected_water_level_count = 0

    # Apply correction only to rows where 'Water Level (m)' is missing
    if 'Water Level (m)' in df.columns:
        missing_data_mask = df['Water Level (m)'].isna()

        # If there are rows with missing data
        if missing_data_mask.any():
            for idx, row in df[missing_data_mask].iterrows():
                # Calculate the water level for missing values
                if pd.isna(row['Water Level (m)']) and pd.notna(row['Differential Pressure (kPa)']):
                    T = row['Temperature (°C)']
                    rho = 999.84 - 0.067 * T  # Simplified density formula for freshwater (kg/m³)
                    df.at[idx, 'Water Level (m)'] = (row['Differential Pressure (kPa)'] * 1000) / (rho * g)
                    corrected_water_level_count += 1

    return df, corrected_water_level_count

def save_formatted_excel(df, output_path):
    """
    Save DataFrame to Excel with standard formatting.
    
    Args:
        df (pd.DataFrame): DataFrame to save
        output_path (str or Path): Path to save the Excel file
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write the DataFrame to Excel
        df.to_excel(writer, sheet_name='Data')
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # Auto-adjust column widths based on content
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add a little extra width and set the column width
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def autodetect_stage_site(file):
    """Detects the stage site based on the file name."""
    if ("22084122" in file.name) or ("cat_beaconsBT" in file.name):
        return "cat_beaconsBT"
    elif ("22084123" in file.name) or ("northfield_poolBT" in file.name):
        return "northfield_poolBT"
    elif ("22084124" in file.name) or ("chase_usBT" in file.name):
        return "chase_usBT"
    elif "cat_beacons_" in file.name:
        return "cat_beacons"
    elif "northfield_bridgeBT" in file.name:
        return "northfield_bridgeBT"
    elif "northfield_bridge" in file.name:
        return "northfield_bridge"
    elif ("chase_us" in file.name) or ("chase_upstream" in file.name):
        return "chase_us"
    elif ("chase_ds" in file.name) or ("chase_downstream" in file.name):
        return "chase_ds"
    else:
        return None  # Default case if no match is found
    
def get_salt_dump_times(site_name):
    # Mapping of input site_name to sheet's Site_Name values
    site_name_mapping = {
        'northfield': 'Northfield',
        'chase_bridge': 'Chase Bridge',
        'cat_beacons': 'Cat Creek (Beaconsfield)'
    }

    # Check if the input site_name is valid
    if site_name not in site_name_mapping:
        raise ValueError(f"Invalid site name: {site_name}. Expected one of 'northfield', 'chase bridge', or 'cat_beacons'.")

    # Get the corresponding sheet Site_Name
    mapped_site_name = site_name_mapping[site_name]

    # Connect to Google Sheets
    gc = gspread.service_account_from_dict(credentials)

    # Open the Google Sheet by URL
    sheet_url = "https://docs.google.com/spreadsheets/d/1JLbDJq4qAfAyzEpOuxYYjhfXd4FxvotUc8JaBCsSdKE/edit?gid=748389405"
    sh = gc.open_by_url(sheet_url)

    # List to hold all salt dump times
    salt_dump_times = []

    # Read all worksheets into memory at once (optimized)
    worksheets = sh.worksheets()

    for ws in worksheets:
        # Read worksheet into DataFrame
        df_ws = pd.DataFrame(ws.get_all_records())

        # Filter rows where 'Site_Name' matches the mapped site_name
        filtered_df = df_ws[df_ws['Site_Name'] == mapped_site_name]

        # Filter out rows where 'Salt_Dump.Time_of_Salt_Dump' is NaN or empty
        valid_salt_dump_times = filtered_df['Salt_Dump.Time_of_Salt_Dump'].dropna()

        # Convert the valid 'Salt_Dump.Time_of_Salt_Dump' values to datetime
        salt_dump_times.extend(pd.to_datetime(valid_salt_dump_times, errors='coerce'))

    return salt_dump_times


    